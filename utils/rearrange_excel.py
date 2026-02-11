import openpyxl

def rearrange_excel_columns(file_path):
    print(f"Rearranging columns in {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Current headers:
    # 0: Faktur Pajak Code
    # 1: Invoice Number
    # 2: Pengusaha Kena Pajak
    # 3: Status
    # 4: Status Invoices
    
    # We want to move index 4 to index 2 (Column C)
    
    # Get all data
    data = []
    for row in sheet.iter_rows(values_only=True):
        row_list = list(row)
        # Move element at index 4 to index 2
        status_inv = row_list.pop(4)
        row_list.insert(2, status_inv)
        data.append(row_list)
    
    # Clear the sheet
    sheet.delete_rows(1, sheet.max_row)
    
    # Write back the rearranged data
    for row_data in data:
        sheet.append(row_data)
        
    wb.save(file_path)
    print("Columns rearranged successfully.")
    # Show new headers
    new_wb = openpyxl.load_workbook(file_path)
    new_sheet = new_wb.active
    print(f"New headers: {[cell.value for cell in new_sheet[1]]}")

if __name__ == "__main__":
    rearrange_excel_columns("faktur pajak.xlsx")
