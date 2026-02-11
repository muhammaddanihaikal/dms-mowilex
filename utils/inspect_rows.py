from openpyxl import load_workbook

def check_rows():
    wb = load_workbook('faktur pajak.xlsx')
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    
    # Header is row 0 in this list
    print(f"Header: {rows[0]}")
    
    # Row 25 (Index 24 in 0-based list including header)
    # Wait, iter_rows(values_only=True) returns a list.
    # rows[0] = Header (row 1)
    # rows[1] = Data Row 1 (row 2)
    # rows[24] = Data Row 24 (row 25)
    
    # My test code loop: 
    # data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    # data_rows[0] = Data Row 1 (row 2)
    # data_rows[23] = Data Row 24 (row 25)
    
    indices_to_check = [24, 25] # Excel rows 25 and 26 correspond to indices 24 and 25 in the full list (including header)
    
    for i in indices_to_check:
        print(f"Row {i+1}: {rows[i]}")

if __name__ == "__main__":
    check_rows()
