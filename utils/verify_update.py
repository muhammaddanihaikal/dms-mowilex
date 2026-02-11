from openpyxl import load_workbook

def verify_update():
    wb = load_workbook('faktur pajak.xlsx')
    sheet = wb.active
    # Row 30 and 31 (indices 29 and 30 in 0-based list)
    rows = list(sheet.iter_rows(values_only=True))
    
    print(f"Row 30: {rows[29]}")
    print(f"Row 31: {rows[30]}")

if __name__ == "__main__":
    verify_update()
