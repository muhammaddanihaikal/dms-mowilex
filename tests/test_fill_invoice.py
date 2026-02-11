"""
TEST: Sinkronisasi data antara file Excel dan aplikasi DMS.
Mengisi nomor invoice yang kosong di Excel dari Aplikasi atau PDF, dan sebaliknya.
"""
import pytest
import os
from dotenv import load_dotenv
import openpyxl
from pages.login_page import LoginPage
from pages.faktur_pajak_page import FakturPajakPage

load_dotenv()

def test_scrape_invoices_to_excel(page):
    # 1. Load Excel Data
    input_file = 'faktur pajak.xlsx'

    
    print(f"Loading {input_file}...")
    try:
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
        # Convert iterator to list to access by index
        # We need the actual row objects to write back to them
        # sheet.rows is a generator
        all_rows = list(sheet.rows)
        # Skip header (row 0)
        data_rows = all_rows[1:] 
        print(f"Loaded {len(data_rows)} data rows.")
    except Exception as e:
        pytest.fail(f"Could not read Excel file: {e}")

    # 2. Login
    login_page = LoginPage(page)
    login_page.navigate()
    
    username = os.getenv("DMS_USERNAME")
    password = os.getenv("DMS_PASSWORD")
    
    if not username or not password:
        pytest.fail("Credentials not found in .env file")
        
    login_page.login(username, password)
    
    # 3. Navigate to Faktur Pajak
    faktur_pajak_page = FakturPajakPage(page)
    faktur_pajak_page.navigate_to_menu()
    
    # 4. Process Rows
    # We will iterate and fill in Invoice Number if it found in App
    
    # indices mapping based on previous inspections:
    # Col 0: Faktur Pajak Code
    # Col 1: Invoice Number
    
    records_updated = 0
    records_processed = 0
    
    # Process only a subset for safety/testing if needed, or all.
    # User said "lanjutkan" imply continuing the task.
    # Let's process rows 25-26 again as a test, or all of them.
    # But usually we want to process all missing ones.
    
    # Finding columns index dynamically is better but hardcoded is fine for now based on previous check
    col_fp_code = 0
    col_invoice = 1
    col_status_inv = 2
    col_file_size = 6 # Based on inspect_rows.py output
    
    print("Starting sync process (App -> Excel)...")
    
    # Let's iterate all, or a specific range if requested. 
    # User previously asked for rows 25-26. I will stick to a small batch or logic to skip already filled?
    # Better to iterate all (or limit to a reasonable number like 50 for a test run)
    
    # --- CONFIGURATION ---
    # Ganti baris ini untuk test range tertentu (Nomor baris di Excel)
    START_ROW = 300
    END_ROW = 499
    # ---------------------
    
    # Calculate slice indices
    start_idx = START_ROW - 2
    
    if END_ROW == -1:
        end_idx = len(data_rows)
        print(f"Testing rows {START_ROW} to END ({len(data_rows)+1})...")
    else:
        end_idx = END_ROW - 2 + 1
        print(f"Testing rows {START_ROW} to {END_ROW}...")
    
    # Ensure indices are valid
    if start_idx < 0: start_idx = 0
    
    target_rows = data_rows[start_idx:end_idx]
    
    for i, row in enumerate(target_rows):
        actual_row_num = i + START_ROW 
        cell_fp = row[col_fp_code]
        cell_inv = row[col_invoice]
        cell_status_excel = row[col_status_inv]
        
        fp_code = str(cell_fp.value).strip() if cell_fp.value else None
        current_excel_inv = str(cell_inv.value).strip() if cell_inv.value and str(cell_inv.value).strip() != "Not extracted" else None
        
        if not fp_code:
            continue
            
        print(f"[{actual_row_num}] Searching for {fp_code}...")
        
        # Search in App
        found = faktur_pajak_page.search(fp_code)
        
        if found:
            rows = faktur_pajak_page.get_all_rows()
            if rows:
                target_row = rows[0]
                found_code = faktur_pajak_page.get_faktur_pajak_code(target_row)
                
                if found_code == fp_code:
                    app_invoice = faktur_pajak_page.get_invoice_number(target_row)
                    app_status = faktur_pajak_page.get_status_invoice(target_row)
                    
                    invoice_to_use = current_excel_inv
                    
                    # CASE A: Excel missing invoice
                    if not invoice_to_use:
                        # Check if App already has a valid-looking invoice
                        if app_invoice and app_invoice not in ["-", "Not extracted"]:
                            print(f"    App already has invoice: '{app_invoice}'. Taking it.")
                            invoice_to_use = app_invoice
                            records_updated += 1
                        else:
                            print(f"    Excel & App missing invoice. Extracting from PDF fallback...")
                            pdf_path = faktur_pajak_page.download_pdf(target_row)
                            if pdf_path:
                                extracted_inv = faktur_pajak_page.extract_invoice_from_pdf(pdf_path)
                                print(f"    Extracted: '{extracted_inv}'")
                                
                                invoice_to_use = extracted_inv if extracted_inv else "Not extracted"
                                records_updated += 1
                            else:
                                print("    Failed to download PDF.")
                                invoice_to_use = "Not extracted"
                    
                    # CASE B: Update App if status not Match
                    if app_status != "Match" and invoice_to_use and invoice_to_use != "Not extracted":
                        print(f"    App status is '{app_status}'. Updating app with '{invoice_to_use}'...")
                        faktur_pajak_page.update_invoice_number(target_row, invoice_to_use)
                        # Refresh data
                        app_status = faktur_pajak_page.get_status_invoice(target_row)
                    
                    # 3. Always write updated values to Excel
                    cell_inv.value = invoice_to_use
                    cell_status_excel.value = app_status
                    print(f"    -> Synced: Inv='{invoice_to_use}', Status='{app_status}'")
                else:
                    print(f"    Mismatch in search result: searched {fp_code}, found {found_code}")
            else:
                print("    Search returned no rows.")
        else:
             print("    Not found in App.")
             
    print(f"Finished processing. Saving to {input_file}...")
    wb.save(input_file)
    print(f"Saved to {input_file}. Total records updated: {records_updated}")
    
    # VERIFICATION AFTER SAVE
    print("Verification of file content on disk:")
    try:
        verify_wb = openpyxl.load_workbook(input_file, data_only=True)
        verify_sheet = verify_wb.active
        # Check row 300 (one of the rows we just updated)
        v_row = 300
        v_code = verify_sheet.cell(row=v_row, column=1).value
        v_inv = verify_sheet.cell(row=v_row, column=2).value
        v_status = verify_sheet.cell(row=v_row, column=3).value
        print(f"FILE ON DISK [Row {v_row}]: Code={v_code}, Inv={v_inv}, Status={v_status}")
    except Exception as e:
        print(f"Could not verify file: {e}")
